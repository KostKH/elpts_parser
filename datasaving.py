"""Модуль парсера eltps_parser, в который вынесены функции для сохранения
результатов в excel-файл"""
import openpyxl


def save_titles_to_excel(filename):
    """Функция создает файл excel, если его нет. Если файл есть - заменяет
    в нём лист elpts на новый (а старый вместе со всеми данными удаляет).
    И в конце записывает первую строку с назавниями колонок. На вход
    принимает имя файла."""
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    if 'elpts' in workbook.sheetnames:
        del workbook['elpts']
    worksheet = workbook.create_sheet(title='elpts')
    titles = [
        'Регистрационный номер свидетельства',
        'Марка транспортного средства',
        'Тип транспортного средства',
        'Статус',
        'Дата оформления',
        'Наименование заявителя',
        'Наименование изготовителя',
        'Наименование испытательной лаборатории',
        'Экологический класс',
        'Номер БСО титульного листа',
        'Номер БСО доп. листов',
        'Печатная форма свидетельства'
    ]
    worksheet.append(titles)
    workbook.save(filename)
    workbook.close()


def save_data_to_excel(filename, results):
    """Функция сохраняет в файл excel спарсенные строки. На вход принимает
    имя файла и список со спарсенными строками."""
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook['elpts']
    for line in results:
        worksheet.append(line)
    workbook.save(filename)
    workbook.close()


def prepare_data(data_dict, result_list):
    """Функция преобразует каждую запись словаря (вытащенного из json
    в результате парсинга страницы) в формат, нужный для последующей
    вставки в excel-файл. После этого запись добавляется в список
    результатов. На вход принимает словарь и список результатов."""
    for item in data_dict:
        data = [
            item['number'],
            item['mark'],
            item['type'],
            item['status'],
            item['activeDate'],
            item['applicant'],
            item['manufacturer'],
            item['sert_org'],
            item['ecoClass'],
            item['blank'],
            item['additionalSheetBlank'],
            f'https://portal.elpts.ru/ncher/sbktsupload?num={item["number"]}',
        ]
        result_list.append(data)
